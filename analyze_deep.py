import pandas as pd
import openpyxl
import json

file_path = 'Project_X_Mobilisation_Plan.xlsx'

def analyze_excel(path):
    workbook = openpyxl.load_workbook(path, data_only=False)
    report = {"sheets": {}}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        df = pd.read_excel(path, sheet_name=sheet_name, header=None) # Read without header to find the real one
        
        formulas = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas[cell.coordinate] = cell.value

        report["sheets"][sheet_name] = {
            "all_data": df.fillna("").values.tolist()[:100], # First 100 rows
            "formulas": formulas
        }

    with open('excel_deep_analysis.json', 'w') as f:
        json.dump(report, f, indent=4, default=str)

if __name__ == "__main__":
    analyze_excel(file_path)
    print("Deep analysis complete.")
