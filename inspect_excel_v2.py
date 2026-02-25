import openpyxl
import json
from datetime import datetime

def json_serial(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    return str(obj)

def inspect_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    report = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = []
        for row in sheet.iter_rows(max_row=50, values_only=True):
            data.append([json_serial(cell) if cell is not None else None for cell in row])
        report[sheet_name] = {
            "dimensions": f"{sheet.max_row}x{sheet.max_column}",
            "sample_data": data
        }
    print(json.dumps(report, indent=2))

if __name__ == "__main__":
    inspect_excel("/Users/mindset/Desktop/Project/Mindset/Project_X_Mobilisation_Plan.xlsx")
