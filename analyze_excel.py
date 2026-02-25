import pandas as pd
import openpyxl
import json
import os

file_path = 'Project_X_Mobilisation_Plan.xlsx'

def analyze_excel(path):
    workbook = openpyxl.load_workbook(path, data_only=False)
    sheet_names = workbook.sheetnames
    report = {
        "sheets": {}
    }

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        sheet_data = {
            "columns": [],
            "formulas": {},
            "sample_data": []
        }

        # Use pandas for easier data extraction
        df = pd.read_excel(path, sheet_name=sheet_name)
        sheet_data["columns"] = df.columns.tolist()
        sheet_data["sample_data"] = df.head(5).to_dict(orient='records')

        # Extract formulas
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    sheet_data["formulas"][cell.coordinate] = cell.value

        report["sheets"][sheet_name] = sheet_data

    with open('excel_analysis.json', 'w') as f:
        json.dump(report, f, indent=4, default=str)

if __name__ == "__main__":
    analyze_excel(file_path)
    print("Analysis complete. Results in excel_analysis.json")
